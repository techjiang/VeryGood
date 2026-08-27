"""xlsx 抽取：逐 sheet 导出 TSV（按显示优先：有缓存值用值，否则回退公式文本）+ 超链接。

用 openpyxl 的 read_only 模式 + iter_rows 流式读取：普通 load_workbook 会把整表解析成
Cell 对象，两个视图（值 / 公式）各需一次全量解析（3 万格实测各约 0.5s）；read_only 是惰性
解析，同规模两视图合计约 0.6s，快一倍。**注意 read_only 下不能用 ws.cell(row, col) 随机
访问**——每次调用都会重新流式扫一遍工作表（实测 2000 格 18.5s），只能按行迭代。
read_only 下 max_row/max_column 取自 XML 的 dimension 元素，第三方写入方可能缺失或写错，
故行列数一律边迭代边计数，不依赖它们（顺带修掉"有格式的空行"把行数虚报的问题）。
"""
from itertools import zip_longest

from .common import merge_links, ooxml_external_links, text_urls

MAX_ROWS = 500      # 每 sheet 抽取行数上限
MAX_COLS = 50       # 每行列数上限


def _cell_str(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\t", " ").strip()


def _close(*wbs):
    """read_only 模式持有 zip 文件句柄，必须显式关闭；None 表示该视图未打开成功。"""
    for wb in wbs:
        if wb is None:
            continue
        try:
            wb.close()
        except Exception:
            pass


def extract(path):
    units, outline, notes = [], [], []
    try:
        import openpyxl
    except Exception:
        return {
            "units": [],
            "links": merge_links(ooxml_external_links(path)),
            "outline": [],
            "meta": {},
            "notes": ["openpyxl 不可用，xlsx 内容未抽取（pip install openpyxl）"],
        }
    wb_v = wb_f = None
    try:
        wb_v = openpyxl.load_workbook(path, read_only=True, data_only=True)   # 缓存值视图
        wb_f = openpyxl.load_workbook(path, read_only=True, data_only=False)  # 公式文本视图
    except Exception as e:
        _close(wb_v, wb_f)  # 公式视图打开失败时缓存值视图已持有 zip 句柄，返回前须关掉
        return {"units": [], "links": [], "outline": [], "meta": {}, "notes": [f"xlsx 打开失败: {e}"]}

    text_links = []
    try:
        sheets = wb_v.sheetnames
        for title in sheets:
            ws_v = wb_v[title]
            ws_f = wb_f[title] if title in wb_f.sheetnames else None
            rows_v = ws_v.iter_rows(values_only=True)
            rows_f = ws_f.iter_rows(values_only=True) if ws_f is not None else iter(())
            rows_seen, rows_out, cols_seen, row_capped, col_capped = 0, 0, 0, False, False
            for row_v, row_f in zip_longest(rows_v, rows_f, fillvalue=()):
                row_v, row_f = row_v or (), row_f or ()
                rows_seen += 1
                if rows_seen > MAX_ROWS:
                    row_capped = True
                    break
                width = max(len(row_v), len(row_f))
                cols_seen = max(cols_seen, width)
                if width > MAX_COLS:
                    col_capped = True
                cells = []
                for c in range(min(width, MAX_COLS)):
                    v = _cell_str(row_v[c] if c < len(row_v) else None)
                    if not v:  # 无缓存值时回退公式文本
                        v = _cell_str(row_f[c] if c < len(row_f) else None)
                    cells.append(v)
                if not any(cells):
                    continue
                line = "\t".join(cells).rstrip("\t")
                units.append({"anchor": f"xlsx:{title}!R{rows_seen}", "text": line})
                text_links += text_urls(line, f"{title}!R{rows_seen}")
                rows_out += 1
            if row_capped:
                notes.append(f"{title} 超过 {MAX_ROWS} 行，仅抽取前 {MAX_ROWS} 行（总行数未统计）")
            if col_capped:
                notes.append(f"{title} 共 {cols_seen} 列，仅抽取前 {MAX_COLS} 列")
            scanned = min(rows_seen, MAX_ROWS)
            outline.append(
                {
                    "anchor": f"xlsx:{title}",
                    "label": f"{title}（已扫 {scanned}{'+' if row_capped else ''} 行 × {cols_seen} 列，非空行 {rows_out}）",
                    "level": 1,
                }
            )
            # 单元格超链接的 URL 只存在于 sheet 的 rels（cell.hyperlink 只带 rId），
            # 故统一由 ooxml_external_links 抽取；位置粒度到 sheet 级（sheetN），不到单元格。
    finally:
        _close(wb_v, wb_f)

    return {
        "units": units,
        "links": merge_links(ooxml_external_links(path), text_links),
        "outline": outline,
        "meta": {"sheets": sheets},
        "notes": notes,
    }
